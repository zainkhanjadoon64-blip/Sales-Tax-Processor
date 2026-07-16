"""
Section 165 Withholding Tax Statement service.

Handles PDF extraction, Excel building/reading, and record normalization.
Ported from the Next.js withholding-tax-statement project.
"""
import io
import os
import re
import shutil
import tempfile
import uuid
from datetime import datetime
from typing import Any

import openpyxl


from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── PDF Extraction ────────────────────────────────────────────────────────────


async def extract_pdf_lines(buffer: bytes) -> list[str]:
    """Reconstruct text lines from PDF using pymupdf."""
    try:
        import fitz  # pymupdf
    except ImportError:
        raise ImportError("pymupdf (fitz) is required for PDF extraction")

    doc = fitz.open(stream=buffer, filetype="pdf")
    page = doc[0]
    blocks = page.get_text("blocks")
    doc.close()

    # Sort by y-coordinate (top to bottom), group within 2px tolerance
    rows: dict[int, list[tuple[float, str]]] = {}
    for b in blocks:
        y = round(b[1])  # y0
        text = (b[4] or "").strip()
        if not text:
            continue
        x = b[0]
        # Find within tolerance
        key = next((k for k in rows if abs(k - y) <= 2), y)
        rows.setdefault(key, []).append((x, text))

    sorted_ys = sorted(rows.keys(), reverse=False)
    lines: list[str] = []
    for y in sorted_ys:
        items = sorted(rows[y], key=lambda it: it[0])
        line = " ".join(it[1] for it in items)
        line = re.sub(r"\s+", " ", line).strip()
        if line:
            lines.append(line)
    return lines


def extract_payment_code(text: str) -> str | None:
    m = re.search(r"Payment\s+Section\s*:[^\n]*?-\s*(\d{8})", text, re.I)
    if m:
        return m.group(1)
    fallback = re.search(r"\b(64\d{6})\b", text)
    return fallback.group(1) if fallback else None


def extract_generation_date(text: str) -> str | None:
    m = re.search(r"Generation\s+Date\s*:\s*(\d{1,2}-[A-Za-z]{3}-\d{4})", text, re.I)
    return m.group(1) if m else None


def parse_amount_pair(segment: str) -> tuple[float, float] | None:
    nums = re.findall(r"[\d,]+(?:\.\d+)?", segment)
    if len(nums) < 2:
        return None
    amount = float(nums[0].replace(",", ""))
    tax = float(nums[1].replace(",", ""))
    return (amount, tax)


LOCATION_CODES = re.compile(r"^[A-Z]{2,3}$")
ADDRESS_KEYWORDS = re.compile(
    r"\b(OPP|PLOT|HOUSE|STREET|ROAD|BLOCK|SECTOR|NEAR|SHOP|FLOOR|MOHALLA|BAZAR|MARKET|CHOWK|COLONY)\b", re.I
)
COMPANY_KEYWORDS = re.compile(
    r"\b(MEDICAL|STORE|LIMITED|LTD|PVT|PRIVATE|TRADERS|ENTERPRISES|PHARMACY|AGENCY|COMPANY|"
    r"CORPORATION|INDUSTRIES|MILLS|BROTHERS|SONS)\b", re.I
)


def extract_single_entry(lines: list[str], start_idx: int, serial_num: int) -> dict | None:
    line = lines[start_idx]
    line = re.sub(rf"^{serial_num}\s+", "", line)

    # Format 2 (Section 236H): "... NAME 236H / B01131 AMOUNT TAX"
    f2 = re.match(r"^(.*?)\s+236H\s*/\s*B\d+\s+(.*)$", line, re.I)
    if f2:
        name_part = f2.group(1)
        name_part = re.sub(r"^(LTO|RTO|CTO)\b[^A-Z]*(\/\s*)?", "", name_part, flags=re.I)
        name_part = re.sub(r"^/\s*", "", name_part).strip()
        amounts = parse_amount_pair(f2.group(2))
        ntn_cnic = ""
        for j in range(start_idx + 1, min(start_idx + 5, len(lines))):
            cm = re.search(r"\b(\d{5}-\d{7}-\d|\d{13})\b", lines[j])
            if cm:
                ntn_cnic = cm.group(1)
                break
        for j in range(start_idx + 1, min(start_idx + 5, len(lines))):
            cand = lines[j].strip()
            if not cand:
                continue
            if re.match(r"^\d", cand):
                break
            if LOCATION_CODES.match(cand):
                continue
            if ADDRESS_KEYWORDS.search(cand):
                continue
            if COMPANY_KEYWORDS.search(cand):
                existing_words = set(name_part.upper().split())
                extra = " ".join(
                    w for w in cand.split() if w.upper() not in existing_words
                )
                if extra:
                    name_part = f"{name_part} {extra}".strip()
                break
        if name_part and amounts:
            return {
                "serial": serial_num,
                "ntn_cnic": ntn_cnic,
                "business_name": name_part,
                "payment_code": "",
                "amount": amounts[0],
                "tax_amount": amounts[1],
            }
        return None

    # Format 1 (Section 153): "NTN OFFICE / NAME 153(1)(a) / AMOUNT TAX"
    ntn_match = re.match(r"^(\d{7}-\d)\s+(.*)$", line)
    if ntn_match:
        ntn = ntn_match.group(1)
        rest = ntn_match.group(2)
        rest = re.sub(r"^(LTO|RTO|CTO)\b[^/]*/\s*", "", rest, flags=re.I).strip()
        sec_match = re.match(r"^(.*?)\s+153\(1\)\([a-z]\)\s*/\s*(.*)$", rest, re.I)
        business_name = ""
        amounts = None
        if sec_match:
            business_name = sec_match.group(1).strip()
            amounts = parse_amount_pair(sec_match.group(2))
        else:
            tail = re.match(r"^(.*?)\s+([\d,]+(?:\.\d+)?)\s+([\d,]+(?:\.\d+)?)$", rest)
            if tail:
                business_name = re.sub(
                    r"\s+\d{3}\(\d\)[a-z]\)\s*/?\s*$", "", tail.group(1), flags=re.I
                ).strip()
                amounts = parse_amount_pair(f"{tail.group(2)} {tail.group(3)}")
        payment_code = ""
        for j in range(start_idx + 1, min(start_idx + 4, len(lines))):
            pc = re.search(r"\b(64\d{6})\b", lines[j])
            if pc:
                payment_code = pc.group(1)
                break
        if business_name and amounts:
            return {
                "serial": serial_num,
                "ntn_cnic": ntn,
                "business_name": business_name,
                "payment_code": payment_code,
                "amount": amounts[0],
                "tax_amount": amounts[1],
            }
    return None


def extract_all_entries(lines: list[str]) -> list[dict]:
    entries: list[dict] = []
    section_start = 0
    for i, line in enumerate(lines):
        if re.search(r"Details\s+of\s+Tax\s*Payers", line, re.I):
            section_start = i + 1
            break
    expected_serial = 1
    header_tokens = ["Sr.", "NTN / CNIC", "Address", "/ NAM Code"]
    for i in range(section_start, len(lines)):
        line = lines[i]
        if any(t in line for t in header_tokens):
            continue
        serial_match = re.match(r"^(\d{1,4})\s+", line)
        if not serial_match:
            continue
        serial = int(serial_match.group(1))
        if serial != expected_serial:
            continue
        entry = extract_single_entry(lines, i, serial)
        if entry:
            entries.append(entry)
            expected_serial += 1
    return entries


async def extract_pdf_data(buffer: bytes) -> dict:
    """Extract data from a challan PDF. Returns dict with success, entries, etc."""
    try:
        lines = await extract_pdf_lines(buffer)
        full_text = "\n".join(lines)
        payment_code = extract_payment_code(full_text)
        generation_date = extract_generation_date(full_text)
        entries = extract_all_entries(lines)
        for e in entries:
            if not e.get("payment_code") and payment_code:
                e["payment_code"] = payment_code
        if not entries:
            return {
                "success": False,
                "error": "No taxpayer entries could be extracted from this PDF",
                "entries": [],
                "payment_code": payment_code,
                "generation_date": generation_date,
            }
        return {
            "success": True,
            "entries": entries,
            "payment_code": payment_code,
            "generation_date": generation_date,
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "entries": [],
            "payment_code": None,
            "generation_date": None,
        }


# ── Formatting ────────────────────────────────────────────────────────────────


def format_ntn_cnic_for_storage(value: str | None) -> str:
    if not value:
        return ""
    value = str(value).strip()
    if "-" in value:
        parts = value.split("-")
        if len(parts) == 3:
            return "".join(parts)  # CNIC: 13101-12184639-1 → 13101121846391
        if len(parts) == 2:
            return parts[0]  # NTN: 0711631-4 → 0711631 (dash + check digit hatao)
    return value


def make_session_id() -> str:
    return f"stmt_{int(datetime.utcnow().timestamp() * 1000)}_{uuid.uuid4().hex[:6]}"


# ── Excel Building (official FBR .xlsm template) ──────────────────────────────


# Project root: services -> app -> backend -> project root
_PROJECT_ROOT = os.path.dirname(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
)
TEMPLATE_PATH = os.path.join(
    _PROJECT_ROOT,
    "Section 165 Statement Template",
    "Withholding_Tax_Statement.xlsm",
)
TEMPLATE_SHEET = "Withholding Data"
DATA_START_ROW = 4  # headers are on row 3; data begins on row 4
VALIDATE_MACRO = "validateWFData"


def _write_records_to_sheet(ws, records: list[dict]) -> None:
    """Write normalized records into the template sheet starting at row 4 using openpyxl."""
    TEXT_FORMAT = "@"
    for idx, r in enumerate(records):
        row = DATA_START_ROW + idx
        c1 = ws.cell(row=row, column=1, value=str(r.get("registration_no", "") or ""))
        c1.number_format = TEXT_FORMAT
        c2 = ws.cell(row=row, column=2, value=str(r.get("identification_no", "") or ""))
        c2.number_format = TEXT_FORMAT
        c3 = ws.cell(row=row, column=3, value=str(r.get("name", "") or ""))
        c3.number_format = TEXT_FORMAT
        c4 = ws.cell(row=row, column=4, value=str(r.get("transaction_date", "") or ""))
        c4.number_format = TEXT_FORMAT
        c5 = ws.cell(row=row, column=5, value=str(r.get("payment_code", "") or "").strip())
        c5.number_format = TEXT_FORMAT
        ws.cell(row=row, column=6, value=r.get("taxable_amount", 0) or 0)
        c7 = ws.cell(row=row, column=7, value=str(r.get("exemption_code", "") or ""))
        c7.number_format = TEXT_FORMAT
        ws.cell(row=row, column=8, value=r.get("tax_amount", 0) or 0)


def _build_via_com(records: list[dict]) -> bytes:
    """Deprecated — use build_statement_workbook instead."""
    return build_statement_workbook(records)


def build_statement_workbook(records: list[dict]) -> bytes:
    """
    Build a Section 165 statement .xlsm workbook using the official FBR template.

    Uses openpyxl to write data (preserves string types), then win32com only
    to trigger the VBA validation macro. Matches the working pattern from
    the FBR Tax Processor project.
    """
    if not os.path.isfile(TEMPLATE_PATH):
        raise FileNotFoundError(
            f"Section 165 template not found at: {TEMPLATE_PATH}. "
            "Cannot generate .xlsm file without the template."
        )

    fd, temp_path = tempfile.mkstemp(suffix=".xlsm")
    os.close(fd)
    try:
        shutil.copy(TEMPLATE_PATH, temp_path)

        # Step 1: Write data with openpyxl (preserves string types, no auto-conversion)
        wb = openpyxl.load_workbook(temp_path, keep_vba=True)
        ws = wb[TEMPLATE_SHEET]
        ws.protection.sheet = False
        _write_records_to_sheet(ws, records)
        wb.save(temp_path)
        wb.close()

        # Step 2: Use COM only for VBA validation macro
        import win32com.client
        import pythoncom

        pythoncom.CoInitialize()
        excel = None
        xl_wb = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False

            xl_wb = excel.Workbooks.Open(os.path.abspath(temp_path))
            ws = xl_wb.Sheets(TEMPLATE_SHEET)

            excel.Application.Run(VALIDATE_MACRO)

            VBA_PASS = "FBRpralIRIS15"
            ws.Unprotect(VBA_PASS)
            last_data_row = DATA_START_ROW + len(records) - 1 if records else DATA_START_ROW
            for r in range(DATA_START_ROW, last_data_row + 1):
                for c in range(1, 10):
                    ws.Cells(r, c).Locked = False
            ws.Protect(VBA_PASS)

            xl_wb.Save()
            xl_wb.Close(False)
            xl_wb = None
            excel.Quit()
            excel = None
        finally:
            if xl_wb is not None:
                try: xl_wb.Close(False)
                except: pass
            if excel is not None:
                try: excel.Quit()
                except: pass
            try: pythoncom.CoUninitialize()
            except: pass

        with open(temp_path, "rb") as f:
            return f.read()
    finally:
        try: os.unlink(temp_path)
        except: pass



# ── Existing Statement Reading ────────────────────────────────────────────────


def _clean_name(raw: str) -> str:
    name = raw.strip()
    # Remove trailing address artifacts: Plot/House/Street followed by number or text
    name = re.sub(r",?\s*(?:Plot|House|Street|St\.?|Shop|Office|Room|Suit)\s*(?:#|No\.?|Number)?\s*\d*[\s,]*.*$", "", name, flags=re.I)
    # Remove standalone bare "Plot Number" or "Plot No" at end (no specific number)
    name = re.sub(r"\s+Plot\s+(?:Number|No\.?)\s*$", "", name, flags=re.I)
    # Remove trailing standalone city names at end
    name = re.sub(r"\s+(LAHORE|KARACHI|ISLAMABAD|RAWALPINDI|ABBOTTABAD|FAISALABAD|MULTAN|GUJRANWALA|PESHAWAR|QUETTA|HYDERABAD|SIALKOT|BAHAWALPUR|SUKKUR|LARKANA)\s*$", "", name, flags=re.I)
    return name.strip()


def read_existing_statement(buffer: bytes) -> list[dict]:
    """Read an existing statement sheet: headers row 3, data from row 4."""
    wb = openpyxl.load_workbook(io.BytesIO(buffer), data_only=True)
    ws = wb.active
    records: list[dict] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        reg_no = str(row[0] or "").strip()
        name = _clean_name(str(row[2] or "").strip())
        if not reg_no and not name:
            continue
        if name.upper() == "TOTAL":
            continue
        records.append({
            "registration_no": reg_no,
            "identification_no": str(row[1] or "").strip(),
            "name": name,
            "transaction_date": str(row[3] or "").strip(),
            "payment_code": str(row[4] or "").strip(),
            "taxable_amount": _parse_float(row[5]),
            "exemption_code": str(row[6] or "").strip(),
            "tax_amount": _parse_float(row[7]),
        })
    wb.close()
    return records


def read_abbottabad_excel(buffer: bytes) -> list[dict]:
    """Read Abbottabad-format import sheet (10 columns)."""
    wb = openpyxl.load_workbook(io.BytesIO(buffer), data_only=True)
    ws = wb.active
    payment_code_map = {"236H/2": "64150801", "236H": "64150801"}
    today = datetime.utcnow().strftime("%Y-%m-%d")
    out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        section = str(row[0] or "").strip()
        ntn = str(row[1] or "").strip()
        cnic = str(row[2] or "").strip()
        name = str(row[3] or "").strip()
        business_name = str(row[7] or "").strip()
        taxable = _parse_float(row[8])
        tax = _parse_float(row[9])
        if not ntn and not cnic and not name:
            continue
        out.append({
            "ntn_cnic": ntn or cnic,
            "business_name": business_name or name,
            "generation_date": today,
            "payment_code": payment_code_map.get(section, "64150801"),
            "amount": taxable,
            "tax_amount": tax,
        })
    wb.close()
    return out


# ── Record Normalization ──────────────────────────────────────────────────────


def normalize_record(r: dict) -> dict:
    """Normalize a frontend/API record (with aliases) into the 8-column template shape."""
    raw_id = (
        str(r.get("ntn_cnic") or "").strip()
        or str(r.get("cnicNtn") or "").strip()  # frontend camelCase
        or str(r.get("cnic") or "").strip()
        or str(r.get("registration_no") or "").strip()
    )
    date_val = (
        str(r.get("generation_date") or "").strip()
        or str(r.get("date") or "").strip()
        or str(r.get("transaction_date") or "").strip()
        or datetime.utcnow().strftime("%d/%m/%Y")
    )
    payment_code = (
        str(r.get("paymentCode") or "").strip()
        or str(r.get("payment_code") or "").strip()
        or str(r.get("code") or "").strip()  # frontend field
    )
    taxable = r.get("taxableAmount", r.get("amount", r.get("taxable_amount", r.get("taxable", 0))))
    tax = r.get("taxAmount", r.get("tax_amount", r.get("tax", 0)))
    return {
        "registration_no": format_ntn_cnic_for_storage(raw_id),
        "identification_no": "",
        "name": str(r.get("business_name") or r.get("name") or "").strip(),
        "transaction_date": date_val,
        "payment_code": payment_code,
        "taxable_amount": _parse_amount(taxable),
        "exemption_code": "",
        "tax_amount": _parse_amount(tax),
    }



def append_to_existing_statement(existing_bytes: bytes, new_records: list[dict]) -> bytes:
    """Append new records to an existing statement .xlsm file.

    Reads the existing file, finds the last data row (>= row 4),
    appends normalized records below, saves, then runs VBA macro.
    Returns the complete .xlsm bytes.
    """
    fd, temp_path = tempfile.mkstemp(suffix=".xlsm")
    os.close(fd)
    try:
        with open(temp_path, "wb") as f:
            f.write(existing_bytes)

        wb = openpyxl.load_workbook(temp_path, keep_vba=True)
        ws = wb[TEMPLATE_SHEET]
        ws.protection.sheet = False

        last_row = DATA_START_ROW - 1
        for row in range(DATA_START_ROW, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val is not None and str(val).strip():
                last_row = row
        start_row = last_row + 1

        TEXT_FORMAT = "@"
        for idx, r in enumerate(new_records):
            row = start_row + idx
            c1 = ws.cell(row=row, column=1, value=str(r.get("registration_no", "") or ""))
            c1.number_format = TEXT_FORMAT
            c2 = ws.cell(row=row, column=2, value=str(r.get("identification_no", "") or ""))
            c2.number_format = TEXT_FORMAT
            c3 = ws.cell(row=row, column=3, value=str(r.get("name", "") or ""))
            c3.number_format = TEXT_FORMAT
            c4 = ws.cell(row=row, column=4, value=str(r.get("transaction_date", "") or ""))
            c4.number_format = TEXT_FORMAT
            c5 = ws.cell(row=row, column=5, value=str(r.get("payment_code", "") or "").strip())
            c5.number_format = TEXT_FORMAT
            ws.cell(row=row, column=6, value=r.get("taxable_amount", 0) or 0)
            c7 = ws.cell(row=row, column=7, value=str(r.get("exemption_code", "") or ""))
            c7.number_format = TEXT_FORMAT
            ws.cell(row=row, column=8, value=r.get("tax_amount", 0) or 0)

        wb.save(temp_path)
        wb.close()

        import win32com.client
        import pythoncom
        pythoncom.CoInitialize()
        excel = None
        xl_wb = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False
            xl_wb = excel.Workbooks.Open(os.path.abspath(temp_path))
            ws = xl_wb.Sheets(TEMPLATE_SHEET)
            excel.Application.Run(VALIDATE_MACRO)
            VBA_PASS = "FBRpralIRIS15"
            ws.Unprotect(VBA_PASS)
            last_new_row = start_row + len(new_records) - 1 if new_records else last_row
            end_row = max(last_row, last_new_row) if new_records else last_row
            for r in range(DATA_START_ROW, end_row + 1):
                for c in range(1, 10):
                    ws.Cells(r, c).Locked = False
            ws.Protect(VBA_PASS)
            xl_wb.Save()
            xl_wb.Close(False)
            xl_wb = None
            excel.Quit()
            excel = None
        finally:
            if xl_wb is not None:
                try: xl_wb.Close(False)
                except: pass
            if excel is not None:
                try: excel.Quit()
                except: pass
            try: pythoncom.CoUninitialize()
            except: pass

        with open(temp_path, "rb") as f:
            return f.read()
    finally:
        try: os.unlink(temp_path)
        except: pass


def _parse_float(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ""))
    except (ValueError, TypeError):
        return 0.0


def _parse_amount(v: Any) -> float:
    n = _parse_float(v)
    return n if n >= 0 else 0.0
